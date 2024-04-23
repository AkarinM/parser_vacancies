import json
import math
import openpyxl
import requests


def parse_vacan(region: str) -> list:
    all_results = list()

    limit = 100  # количество записей на странице

    # Получаем первую страницу
    url = f'https://opendata.trudvsem.ru/api/v1/vacancies/region/{region}'
    result = requests.get(url).json()

    all_results.append(json.dumps(result, indent=2))

    total_pages = math.ceil(result['meta']['total'] / 100)  # Сколько нам нужно получить страниц

    print('total_pages', total_pages)

    # Постранично получаем результаты
    for page in range(1, total_pages):
        url = f'https://opendata.trudvsem.ru/api/v1/vacancies/region/{region}?offset={page}&limit={limit}'
        result = requests.get(url).json()

        all_results.append(json.dumps(result, indent=2))

    return all_results

def write_in_xls(result_list: list):
    wb = openpyxl.Workbook()
    ws = wb.active

    x = 2
    column_names = list()
    for a_r in all_results:
        a_r = json.loads(a_r, encoding='utf-8')
        print(a_r)
        for all_data in a_r['results']['vacancies']:
            all_data = all_data['vacancy']

            data = {
                'job_name': all_data.get('job-name', ''),
                # 'company_dict': all_data.get('company', {}),
                'salary_min': all_data.get('salary_min', 0),
                'salary_max': all_data.get('salary_max', 0),
                'education': all_data.get('requirement', {}).get('education', ''),
                'category': all_data.get('category', {}).get('specialisation', ''),
            }

            company_dict = all_data['company']

            data.update(company_dict)

            if len(column_names) == 0:
                column_names = list(company_dict.keys())
                column_names.append('job_name')
                column_names.append('salary_min')
                column_names.append('salary_max')
                column_names.append('education')
                column_names.append('category')

            for y, name in enumerate(column_names, 1):
                ws.cell(row=x, column=y, value=data.get(name, ''))

            x += 1

    for y, name in enumerate(column_names, 1):
        ws.cell(row=1, column=y, value=name)

    wb.save(f'Вакансии.xls')


if __name__ == '__main__':
    all_results = parse_vacan('7600000000000')
    write_in_xls(all_results)